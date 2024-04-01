import openpyxl

def generate_text_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # 读取数据
    quarter = sheet.cell(row=1, column=2).value
    revenue_this_period = sheet.cell(row=3, column=2).value
    revenue_last_period = sheet.cell(row=3, column=3).value
    revenue_diff = sheet.cell(row=3, column=4).value
    revenue_increase_percentage = sheet.cell(row=3, column=5).value

    # 生成文本
    text = f"{quarter}，我所预计完成税收收入{revenue_this_period}万元，同期完成{revenue_last_period}万元，同比增长{revenue_increase_percentage}%,增收{revenue_diff}万元"
    
    return text

# 示例用法
file_path = "./data.xlsx"
generated_text = generate_text_from_excel(file_path)
print(generated_text)
